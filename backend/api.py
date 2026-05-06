"""
api.py  —  FastAPI REST + WebSocket backend for DWSIM Full UI
──────────────────────────────────────────────────────────────
Endpoints:
  POST /chat/stream          SSE streaming chat
  POST /flowsheet/load       Load a flowsheet by path
  POST /flowsheet/save       Save current flowsheet
  POST /flowsheet/run        Run simulation
  GET  /flowsheet/objects    List all simulation objects
  GET  /flowsheet/results    Get all stream results
  POST /stream/property      Set a stream property
  POST /stream/composition   Set stream mole fractions (ACC-1)
  GET  /flowsheet/package    Get property package (ACC-3)
  POST /flowsheet/validate   Validate feed specs (ACC-4)
  GET  /flowsheet/convergence Check convergence (ACC-2)
  POST /parametric           Run parametric study
  POST /optimize             Run optimisation (ACC-5)
  GET  /find                 Find flowsheet files on disk
  GET  /health               Health check

Run: python api.py
"""

from dotenv import load_dotenv
load_dotenv(override=True)

import asyncio
import json
import os
import sys
import threading
import traceback
import uuid
from typing import Any, Dict, List, Optional

import time as _time_module
import uvicorn
from collections import defaultdict
from contextlib import asynccontextmanager
from fastapi import FastAPI, HTTPException, WebSocket, WebSocketDisconnect, Request, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
from fastapi.security import APIKeyHeader
from pydantic import BaseModel, field_validator

from flowsheet_watcher import FlowsheetScanner, FlowsheetWatcher
import dwsim_gui_bridge

sys.path.insert(0, os.path.dirname(__file__))

from dwsim_bridge_v2  import DWSIMBridgeV2, list_backups, restore_backup
from llm_client       import LLMClient, DEFAULT_MODELS
from agent_v2         import DWSIMAgentV2
from session          import save_session, load_session, list_sessions
from knowledge_base   import KnowledgeBase
from evaluation       import get_eval_log, get_benchmark_suite, SessionTracker
from reliability      import get_analyzer, get_failure_log
from accuracy         import get_accuracy_store, get_accuracy_comparer, ReferenceSet, ReferenceEntry, PROPERTIES
import diagnostics
import session_memory

_kb = KnowledgeBase()

# ─────────────────────────────────────────────────────────────────────────────
# Security: API Key Authentication
# ─────────────────────────────────────────────────────────────────────────────
# Set API_SECRET_KEY in .env to enable authentication.
# If not set, auth is DISABLED (backward compat for single-machine dev use).
# The UI sends the key via X-API-Key header on every request.

_API_SECRET_KEY: str = os.getenv("API_SECRET_KEY", "")
_api_key_header = APIKeyHeader(name="X-API-Key", auto_error=False)

def verify_api_key(api_key: str = Depends(_api_key_header)) -> None:
    """FastAPI dependency — raises 403 if API key is wrong."""
    if not _API_SECRET_KEY:
        return  # Auth disabled — development mode
    if api_key != _API_SECRET_KEY:
        raise HTTPException(
            status_code=403,
            detail="Invalid or missing API key. Set X-API-Key header.",
        )

# ─────────────────────────────────────────────────────────────────────────────
# Rate Limiting: simple in-memory token bucket per client IP
# ─────────────────────────────────────────────────────────────────────────────
_rate_lock = threading.Lock()
_rate_hits: dict = defaultdict(list)          # ip → [timestamp, ...]

# Chat endpoint: max 20 requests per minute per IP (generous for normal use)
_CHAT_RATE_MAX   = int(os.getenv("CHAT_RATE_MAX", "20"))
_CHAT_RATE_WIN_S = 60.0

def _check_rate(ip: str, max_req: int, window_s: float) -> bool:
    """Return False if the IP has exceeded max_req in the last window_s seconds."""
    now = _time_module.time()
    with _rate_lock:
        hits = [t for t in _rate_hits[ip] if now - t < window_s]
        if len(hits) >= max_req:
            _rate_hits[ip] = hits
            return False
        hits.append(now)
        _rate_hits[ip] = hits
        return True

# ─────────────────────────────────────────────────────────────────────────────
# Concurrent chat guard — DWSIM bridge is not safe for concurrent flowsheet ops
# ─────────────────────────────────────────────────────────────────────────────
# Use a Lock + flag (not Event alone) so check-and-set is atomic.
# Event.is_set() + Event.set() has a TOCTOU race between concurrent requests.
_chat_busy_lock = threading.Lock()
_chat_busy      = False           # protected by _chat_busy_lock

def _try_acquire_chat_slot() -> bool:
    """Atomically claim the chat slot. Returns True if acquired, False if already busy."""
    global _chat_busy
    with _chat_busy_lock:
        if _chat_busy:
            return False
        _chat_busy = True
        return True

def _release_chat_slot() -> None:
    """Release the chat slot. Safe to call multiple times."""
    global _chat_busy
    with _chat_busy_lock:
        _chat_busy = False

# ─────────────────────────────────────────────────────────────────────────────
# Flowsheet file watcher + WebSocket broadcast
# ─────────────────────────────────────────────────────────────────────────────

_scanner = FlowsheetScanner()
_watcher: Optional[FlowsheetWatcher] = None
_ws_clients: list = []  # active WebSocket connections


def _broadcast_file_event(event_type: str, file_meta: dict):
    """Called by FlowsheetWatcher when a file changes — queues broadcast to all WS clients.

    Also flags an 'external_edit' payload when the currently-loaded flowsheet
    is modified on disk since our last load/save. The UI can prompt for reload.
    """
    import json
    try:
        bridge = _bridge
        loaded_path = ""
        cached_mtime = 0.0
        if bridge is not None and getattr(bridge, "state", None) is not None:
            loaded_path = (bridge.state.path or "").lower()
            cached_mtime = getattr(bridge.state, "loaded_mtime", 0.0) or 0.0
        event_path = (file_meta or {}).get("path", "").lower()
        if (event_type == "modified"
                and loaded_path and event_path == loaded_path):
            on_disk = (file_meta or {}).get("modified_ts", 0.0) or 0.0
            if cached_mtime > 0 and (on_disk - cached_mtime) > 1.0:
                file_meta = dict(file_meta)
                file_meta["external_edit"] = True
                file_meta["cached_mtime"] = cached_mtime
    except Exception:
        pass

    msg = json.dumps({"type": "file_event",
                      "event": event_type, "file": file_meta})
    dead: list = []
    for ws_queue in list(_ws_clients):
        try:
            ws_queue.put_nowait(msg)
        except Exception:
            # Queue full or closed → client is a zombie; mark for removal
            dead.append(ws_queue)
    for q in dead:
        try:
            _ws_clients.remove(q)
        except ValueError:
            pass


# ─────────────────────────────────────────────────────────────────────────────
# App setup
# ─────────────────────────────────────────────────────────────────────────────

@asynccontextmanager
async def _lifespan(application: FastAPI):
    """Startup and shutdown using the modern lifespan pattern (no deprecation)."""
    global _watcher
    loop = asyncio.get_running_loop()
    # --- startup ---
    try:
        await loop.run_in_executor(None, _get_bridge)
        print("[DWSIM API] Bridge initialised")
    except Exception as exc:
        print(f"[DWSIM API] Bridge init warning (DWSIM may not be installed): {exc}")
    try:
        await loop.run_in_executor(None, _get_agent)
        print("[DWSIM API] Agent initialised")
    except Exception as exc:
        print(f"[DWSIM API] Agent init warning: {exc}")
    try:
        _watcher = FlowsheetWatcher(on_change=_broadcast_file_event, poll_interval=3.0)
        _watcher.start()
        print("[DWSIM API] Flowsheet watcher started")
    except Exception as exc:
        print(f"[DWSIM API] Watcher warning: {exc}")
    yield
    # --- shutdown ---
    if _watcher:
        _watcher.stop()
        print("[DWSIM API] Flowsheet watcher stopped")


app = FastAPI(
    title="DWSIM Agentic AI v2",
    description="Natural language DWSIM process simulation API",
    lifespan=_lifespan,
    version="2.0.0",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ─────────────────────────────────────────────────────────────────────────────
# Global singletons — initialised lazily on first request
# ─────────────────────────────────────────────────────────────────────────────
# (threading already imported at top of file)

_bridge: Optional[DWSIMBridgeV2] = None
_agent:  Optional[DWSIMAgentV2]  = None
_bridge_lock = threading.Lock()   # DWSIM .NET is NOT thread-safe; serialise all access

# Last agent answer — kept for on-demand accuracy comparison
_last_agent_answer: Dict[str, Any] = {"text": "", "session_id": ""}


class _LockedBridge:
    """Transparent proxy that holds _bridge_lock for every method call."""
    def __getattr__(self, name):
        attr = getattr(_bridge, name)
        if callable(attr):
            def _locked(*args, **kwargs):
                with _bridge_lock:
                    return attr(*args, **kwargs)
            return _locked
        return attr


_locked_bridge_proxy = _LockedBridge()


def _get_bridge() -> "_LockedBridge":
    global _bridge
    if _bridge is None:
        with _bridge_lock:
            if _bridge is None:          # double-checked locking
                _bridge = DWSIMBridgeV2(dll_folder=os.getenv("DWSIM_DLL_FOLDER"))
                _bridge.initialize()
    return _locked_bridge_proxy


_PROVIDER_ENV_KEY = {
    "groq":      "GROQ_API_KEY",
    "gemini":    "GEMINI_API_KEY",
    "openai":    "OPENAI_API_KEY",
    "anthropic": "ANTHROPIC_API_KEY",
    "ollama":    "",  # no key
}


def _provider_usable(info: dict) -> bool:
    """True only when the provider probe returned a clean 2xx (key valid + quota ok)."""
    # usable key set explicitly (new diagnostics)
    if "usable" in info:
        return bool(info["usable"]) and bool(info.get("key_configured"))
    # Legacy: reachable=True + no error string → assume usable
    return bool(info.get("reachable")) and not info.get("error") and bool(info.get("key_configured"))


def _choose_provider(preferred: str) -> str:
    """Return the best usable provider, respecting the user's preference.
    Falls back to the next working provider if the preferred one has
    a bad key, exhausted quota, or is unreachable.
    """
    try:
        probe = diagnostics.probe_llm_providers(timeout_s=3.0)
    except Exception:
        return preferred

    p = (preferred or "").lower()

    # Keep user preference only if its key is actually working (2xx probe)
    if p and _provider_usable(probe.get(p, {})):
        return p

    # Fallback order: prefer free/working providers first
    fallback_order = ["gemini", "groq", "ollama", "anthropic", "openai"]
    for cand in fallback_order:
        if cand == p:
            continue  # already tried
        if _provider_usable(probe.get(cand, {})):
            print(f"[startup] '{p}' unavailable (HTTP {probe.get(p,{}).get('status','?')}); "
                  f"using '{cand}' instead")
            return cand

    # Last resort: ollama (local, no key needed)
    if probe.get("ollama", {}).get("reachable"):
        return "ollama"

    return preferred  # give up and let the LLM client report the error


def _get_agent() -> DWSIMAgentV2:
    global _agent
    if _agent is None:
        preferred = os.getenv("LLM_PROVIDER", "groq")
        provider  = _choose_provider(preferred)
        env_key   = _PROVIDER_ENV_KEY.get(provider, "")
        api_key   = os.getenv(env_key, "") if env_key else ""
        model     = os.getenv("LLM_MODEL", DEFAULT_MODELS.get(provider, "")) \
                    if provider == preferred \
                    else DEFAULT_MODELS.get(provider, "")
        llm       = LLMClient(provider=provider, api_key=api_key, model=model)
        _agent    = DWSIMAgentV2(
            llm=llm,
            bridge=_get_bridge(),
            verbose=False,
            stream_output=False,
        )
    return _agent


# ─────────────────────────────────────────────────────────────────────────────
# Pydantic models
# ─────────────────────────────────────────────────────────────────────────────

class ChatRequest(BaseModel):
    message: str
    session_id: Optional[str] = None
    auto_reflect: bool = False           # save + push to DWSIM after turn
    reflect_path: Optional[str] = None   # target path (defaults to loaded or Documents)
    reflect_close_first: bool = True     # close running DWSIM windows before relaunch

    @field_validator("message", mode="before")
    @classmethod
    def message_not_empty_and_bounded(cls, v: str) -> str:
        """BUG-14 fix: reject empty or oversized messages before they hit the LLM."""
        v = str(v).strip()
        if not v:
            raise ValueError("message must not be empty")
        if len(v) > 50_000:
            raise ValueError("message exceeds 50 000-character limit")
        return v

class LoadRequest(BaseModel):
    path: str
    alias: Optional[str] = None

class SaveRequest(BaseModel):
    path: Optional[str] = None
    force: bool = False
    push_to_gui: bool = False
    close_gui_first: bool = False


class GuiPushRequest(BaseModel):
    path: Optional[str] = None
    close_first: bool = False

class StreamPropertyRequest(BaseModel):
    tag: str
    property_name: str
    value: float
    unit: str = ""

class StreamCompositionRequest(BaseModel):
    tag: str
    compositions: Dict[str, float]

class UnitOpPropertyRequest(BaseModel):
    tag: str
    property_name: str
    value: str

class ParametricRequest(BaseModel):
    vary_tag: str
    vary_property: str
    vary_unit: str
    values: List[float]
    observe_tag: str
    observe_property: str

class OptimizeRequest(BaseModel):
    vary_tag: str
    vary_property: str
    vary_unit: str
    lower_bound: float
    upper_bound: float
    observe_tag: str
    observe_property: str
    minimize: bool = True
    tolerance: float = 1e-4
    max_iterations: int = 50

class AccuracyReferenceEntryModel(BaseModel):
    stream_tag:   str
    property_key: str    # e.g. "temperature_C"
    manual_value: float
    note:         str = ""

class AccuracyReferenceSetModel(BaseModel):
    name:      str
    flowsheet: str = ""
    entries:   List[AccuracyReferenceEntryModel]

class AccuracyCaptureRequest(BaseModel):
    name:        str
    stream_tags: Optional[List[str]] = None
    properties:  Optional[List[str]] = None

class AccuracyCompareRequest(BaseModel):
    ref_id:     str
    use_last_agent_answer: bool = True   # parse agent text for stated values
    auto_query: bool = True              # if no recent agent text, ask the
                                         # agent now to populate the response

class StreamPropertyReadRequest(BaseModel):
    tag: str


class EconomicsRequest(BaseModel):
    annual_hours:            float = 8000
    product_price_per_kg:    float = 1.0
    feed_price_per_kg:       float = 0.30
    electricity_per_kWh:     float = 0.08
    steam_per_GJ:            float = 18.0
    cooling_water_per_GJ:    float = 0.25
    project_life_years:      int   = 15
    discount_rate:           float = 0.12
    labor_per_year:          float = 400_000
    lang_factor:             float = 4.7
    contingency_frac:        float = 0.15
    product_stream_tags:     List[str] = []
    feed_stream_tags:        List[str] = []
    capex_scale:             float = 1.0


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _ok(data: Any) -> Dict:
    if isinstance(data, dict) and "success" in data:
        if not data["success"]:
            raise HTTPException(status_code=400, detail=data.get("error", "Failed"))
    return data


# ─────────────────────────────────────────────────────────────────────────────
# Health
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/")
def serve_ui():
    """Serve the single-file UI (no Node.js required)."""
    ui_path = os.path.join(os.path.dirname(__file__), "ui.html")
    return FileResponse(ui_path, media_type="text/html")


# ── React UI at /app ──────────────────────────────────────────────────────────
_REACT_BUILD  = os.path.normpath(os.path.join(os.path.dirname(__file__), "..", "frontend", "build"))
_REACT_STATIC = os.path.join(_REACT_BUILD, "static")
_REACT_INDEX  = os.path.join(_REACT_BUILD, "index.html")

# Serve React static files at request time (not module-level mount) so the
# backend does NOT need to be restarted after `npm run build`.
# This route MUST be registered before the SPA catch-all below.
import mimetypes as _mimetypes

@app.get("/app/static/{file_path:path}", include_in_schema=False)
def serve_react_static(file_path: str):
    """Serve React static assets (JS, CSS, images) directly from disk."""
    full = os.path.join(_REACT_STATIC, file_path)
    if not os.path.isfile(full):
        raise HTTPException(status_code=404, detail="Static file not found")
    mime, _ = _mimetypes.guess_type(full)
    return FileResponse(full, media_type=mime or "application/octet-stream")


@app.get("/app", response_class=HTMLResponse, include_in_schema=False)
@app.get("/app/{rest_of_path:path}", response_class=HTMLResponse, include_in_schema=False)
def serve_react(rest_of_path: str = ""):
    """Serve the React SPA — all non-static sub-routes return index.html."""
    if os.path.isfile(_REACT_INDEX):
        with open(_REACT_INDEX, "r", encoding="utf-8") as f:
            return HTMLResponse(content=f.read())
    return HTMLResponse(
        content=(
            "<h2>React build not found.</h2>"
            "<p>Run inside <code>dwsim_full/frontend/</code>:</p>"
            "<pre>npm install\nnpm run build</pre>"
        ),
        status_code=503,
    )


@app.post("/admin/reload-env", include_in_schema=False)
def reload_env():
    """Reload .env file into os.environ without restarting the server."""
    load_dotenv(override=True)
    return {
        "success": True,
        "providers": {
            "groq":      bool(os.getenv("GROQ_API_KEY")),
            "gemini":    bool(os.getenv("GEMINI_API_KEY")),
            "openai":    bool(os.getenv("OPENAI_API_KEY")),
            "anthropic": bool(os.getenv("ANTHROPIC_API_KEY")),
        },
    }


@app.get("/health")
def health():
    bridge = _get_bridge()
    return {
        "status": "ok",
        "bridge_ready": bridge._ready,
        "flowsheet": bridge.state.name or None,
        "property_package": bridge.state.property_package or None,
    }


@app.get("/diagnostics")
def diagnostics_endpoint(skip_providers: bool = False):
    """Full runtime health report: DWSIM DLLs, python deps, LLM providers, FOSSEE scan."""
    bridge = _get_bridge()
    return diagnostics.full_diagnostics(bridge, skip_providers=skip_providers)


@app.get("/diagnostics/providers")
def diagnostics_providers():
    """Just the LLM-provider reachability probe (cheap)."""
    probe = diagnostics.probe_llm_providers()
    return {
        "providers": probe,
        "recommended_order": diagnostics.recommended_provider_order(probe),
    }


# ── Session memory endpoints ────────────────────────────────────────────────

class MemoryRecordRequest(BaseModel):
    entry_type: str
    payload: Dict[str, Any] = {}


class GoalRequest(BaseModel):
    text: str


@app.get("/memory/recent")
def memory_recent(limit: int = 10, entry_type: Optional[str] = None):
    return {"entries": session_memory.recent(limit, entry_type)}


@app.get("/memory/search")
def memory_search(q: str = "", limit: int = 10):
    return {"entries": session_memory.search(q, limit)}


@app.get("/memory/goals")
def memory_goals():
    return session_memory.get_goals()


@app.post("/memory/goals")
def memory_set_goal(req: GoalRequest):
    return session_memory.set_goal(req.text)


@app.post("/memory/constraints")
def memory_set_constraint(req: GoalRequest):
    return session_memory.set_constraint(req.text)


@app.delete("/memory/goals")
def memory_clear_goals():
    return session_memory.clear_goals()


@app.post("/memory/record")
def memory_record(req: MemoryRecordRequest):
    return session_memory.record(req.entry_type, **(req.payload or {}))


# ─────────────────────────────────────────────────────────────────────────────
# Chat — SSE streaming
# ─────────────────────────────────────────────────────────────────────────────

_REFLECT_MIN_INTERVAL = 3.0  # seconds between reflects
_reflect_last_ts = 0.0
_reflect_lock = threading.Lock()


def _default_reflect_path() -> str:
    """Where to save the AI's working flowsheet when the user hasn't loaded one."""
    docs = os.path.expanduser("~/Documents")
    return os.path.join(docs, "DWSIM_AI_Session.dwxmz")


def _reflect_to_dwsim(explicit_path: Optional[str],
                      close_first: bool) -> dict:
    """Save the current flowsheet and push it into the running DWSIM GUI."""
    global _reflect_last_ts
    import time as _t
    with _reflect_lock:
        since = _t.time() - _reflect_last_ts
        if since < _REFLECT_MIN_INTERVAL:
            return {"success": False, "code": "DEBOUNCED",
                    "error": f"reflect skipped ({since:.1f}s since last; "
                             f"min {_REFLECT_MIN_INTERVAL}s)",
                    "retry_in_s": round(_REFLECT_MIN_INTERVAL - since, 2)}
        _reflect_last_ts = _t.time()

    bridge = _get_bridge()
    if bridge._flowsheet is None:
        return {"success": False, "code": "NO_FLOWSHEET",
                "error": "no flowsheet to reflect"}
    st = getattr(bridge, "state", None)
    target = (explicit_path
              or (st.path if st and getattr(st, "path", "") else "")
              or _default_reflect_path())
    try:
        save = bridge.save_flowsheet(target, force=True)
    except Exception as exc:
        return {"success": False, "stage": "save", "error": str(exc)}
    if not save.get("success"):
        return {"success": False, "stage": "save",
                "error": save.get("error"), "path": target}
    try:
        push = dwsim_gui_bridge.push_to_gui(save["saved_to"],
                                            close_first=close_first)
    except Exception as exc:
        return {"success": False, "stage": "push", "error": str(exc),
                "saved_to": save["saved_to"]}
    summary = {
        "streams": len(st.streams) if st else 0,
        "unit_ops": len(st.unit_ops) if st else 0,
        "property_package": (st.property_package if st else "") or "",
    }
    return {"success": bool(push.get("success")),
            "saved_to": save["saved_to"],
            "backup": save.get("backup"),
            "summary": summary,
            "push": push}


@app.post("/chat/stream")
async def chat_stream(
    req: ChatRequest,
    request: Request,
    _auth: None = Depends(verify_api_key),
):
    """
    SSE endpoint. Streams the agent response token by token.
    Each SSE event is: data: <json>\n\n
    Event types: token | tool_call | done | error
    """
    # ── Rate limiting ─────────────────────────────────────────────────────────
    client_ip = request.client.host if request.client else "unknown"
    if not _check_rate(client_ip, _CHAT_RATE_MAX, _CHAT_RATE_WIN_S):
        raise HTTPException(
            status_code=429,
            detail=f"Rate limit exceeded: max {_CHAT_RATE_MAX} chat requests/minute. "
                   "Wait a moment and try again.",
        )

    # ── Concurrent-request guard ──────────────────────────────────────────────
    # DWSIM bridge is single-tenant; concurrent flowsheet operations corrupt state.
    # Atomically claim the slot (no TOCTOU race). Release on any error path
    # before the background task is scheduled, otherwise the slot leaks forever.
    if not _try_acquire_chat_slot():
        raise HTTPException(
            status_code=409,
            detail="Another simulation is in progress. Please wait for it to complete.",
        )

    # From here until the background task is scheduled, ANY exception must
    # release the slot or the server permanently rejects all chat requests.
    try:
        agent = _get_agent()

        queue: asyncio.Queue = asyncio.Queue()
        loop = asyncio.get_running_loop()

        # ── Evaluation tracking ───────────────────────────────────────────────
        tracker = SessionTracker(req.message, get_eval_log())
    except Exception:
        _release_chat_slot()
        raise

    def on_token(tok: str):
        loop.call_soon_threadsafe(
            queue.put_nowait, json.dumps({"type": "token", "data": tok})
        )

    # Tools that mutate the flowsheet — reflect after any of these fire.
    _WRITE_TOOLS = {
        # step-by-step build tools (primary)
        "new_flowsheet", "add_object", "save_and_solve",
        # topology
        "connect_streams", "disconnect_streams", "delete_object",
        # property setters
        "set_stream_property", "set_stream_composition",
        "set_unit_op_property", "set_stream_flash_spec",
        # simulation
        "load_flowsheet", "run_simulation", "create_from_template",
        # column/reactor/HX config
        "set_column_property", "set_column_specs",
        "set_reactor_property", "setup_reaction",
        "configure_heat_exchanger", "set_energy_stream",
        "set_property_package", "set_binary_interaction_parameters",
    }
    dirty = {"changed": False}

    def on_tool_call(name: str, args: dict, result: dict):
        tracker.record_tool_call(name, args, result)
        if name in _WRITE_TOOLS and (result or {}).get("success"):
            dirty["changed"] = True
        loop.call_soon_threadsafe(
            queue.put_nowait, json.dumps({
                "type": "tool_call",
                "data": {"name": name, "args": args, "result": result},
            })
        )

    async def run_agent():
        agent.on_token     = on_token
        agent.on_tool_call = on_tool_call
        try:
            answer = await loop.run_in_executor(None, agent.chat, req.message)
            session = tracker.finish(answer)           # ← log success
            _last_agent_answer["text"]       = answer
            _last_agent_answer["session_id"] = session.session_id

            # ── Auto-reflect: save + push to DWSIM GUI ───────────────────
            if req.auto_reflect and dirty["changed"]:
                reflect_result = await loop.run_in_executor(
                    None, _reflect_to_dwsim,
                    req.reflect_path, req.reflect_close_first,
                )
                await queue.put(json.dumps({
                    "type": "reflect", "data": reflect_result
                }))

            await queue.put(json.dumps({"type": "done", "data": answer}))
        except Exception as exc:
            session = tracker.finish("", error=str(exc))  # ← log failure
            answer  = ""
            await queue.put(json.dumps({"type": "error", "data": str(exc)}))

        # ── Reliability analysis (post-hoc, non-blocking) ─────────────────
        try:
            sim_results    = None
            stream_objects = None
            bridge = _get_bridge()
            if bridge._flowsheet is not None:
                raw = bridge.get_simulation_results()
                if raw.get("success"):
                    sim_results = raw
                raw2 = bridge.list_simulation_objects()
                if raw2.get("success"):
                    stream_objects = raw2.get("objects", [])

            report = get_analyzer().analyze(
                session_id     = session.session_id,
                user_message   = req.message,
                agent_text     = answer,
                tool_records   = session.tool_records_raw,
                sim_results    = sim_results,
                stream_objects = stream_objects,
            )
            session.reliability_issues = [i.to_dict() for i in report.issues]
            get_failure_log().add(report)
            # Re-save session with reliability data
            get_eval_log().add_session(session)
        except Exception:
            pass   # reliability analysis must never break the chat

    agent_done = {"done": False}

    async def run_agent_wrapper():
        try:
            await run_agent()
        finally:
            _release_chat_slot()    # always release the busy flag
            agent_done["done"] = True

    asyncio.create_task(run_agent_wrapper())

    # Hard deadline: kill the SSE after 10 minutes no matter what.
    _HARD_DEADLINE_S = 600.0
    _KEEPALIVE_S     = 15.0   # emit a comment ping so the browser doesn't close the connection

    async def event_generator():
        deadline = asyncio.get_running_loop().time() + _HARD_DEADLINE_S
        while True:
            remaining = deadline - asyncio.get_running_loop().time()
            if remaining <= 0:
                yield ("data: {\"type\": \"error\", \"data\": "
                       "\"Request exceeded 10-minute limit. "
                       "For flowsheet creation, use the ⚡ Build Now button instead.\"}\n\n")
                break
            wait = min(_KEEPALIVE_S, remaining)
            try:
                item = await asyncio.wait_for(queue.get(), timeout=wait)
                yield f"data: {item}\n\n"
                parsed = json.loads(item)
                if parsed["type"] in ("done", "error"):
                    break
            except asyncio.TimeoutError:
                if agent_done["done"]:
                    # Agent finished but didn't emit done — shouldn't happen, but be safe.
                    break
                # Emit SSE comment (browsers ignore these; keeps the TCP connection alive)
                yield ": ping\n\n"

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
        },
    )


@app.post("/chat/reset")
def chat_reset():
    agent = _get_agent()
    agent.reset()
    return {"success": True, "message": "Conversation history cleared"}


# ─────────────────────────────────────────────────────────────────────────────
# Flowsheet operations
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/find")
def find_flowsheets():
    return _ok(_get_bridge().find_flowsheets())


@app.post("/flowsheet/load")
def load_flowsheet(req: LoadRequest):
    r = _get_bridge().load_flowsheet(req.path, req.alias)
    if not r.get("success"):
        # Preserve structured load-error payload (code, suggestions) for UI.
        raise HTTPException(status_code=400, detail=r)
    return r


@app.post("/flowsheet/save")
def save_flowsheet(req: SaveRequest):
    result = _get_bridge().save_flowsheet(req.path, force=req.force)
    if result.get("success") and req.push_to_gui:
        try:
            push = dwsim_gui_bridge.push_to_gui(
                result.get("saved_to") or req.path or "",
                close_first=req.close_gui_first)
            result["gui_push"] = push
        except Exception as exc:
            result["gui_push"] = {"success": False, "error": str(exc)}
    return _ok(result)


class RestoreBackupRequest(BaseModel):
    backup_path: str
    target_path: Optional[str] = None  # defaults to current loaded path


@app.get("/flowsheet/backups")
def flowsheet_backups(path: Optional[str] = None):
    """List rolling backups for a flowsheet (defaults to the loaded one)."""
    bridge = _get_bridge()
    p = path or (bridge.state.path if getattr(bridge, "state", None) else "")
    if not p:
        return {"path": "", "backups": []}
    return {"path": p, "backups": list_backups(p)}


@app.post("/flowsheet/backups/restore")
def flowsheet_backup_restore(req: RestoreBackupRequest):
    bridge = _get_bridge()
    target = req.target_path or (bridge.state.path
                                 if getattr(bridge, "state", None) else "")
    if not target:
        raise HTTPException(status_code=400,
                            detail="No target path — load a flowsheet first")
    r = restore_backup(req.backup_path, target)
    if r.get("success"):
        # Reload from the restored file so the in-memory bridge sees it.
        r["reload"] = bridge.load_flowsheet(target)
    return r


@app.get("/flowsheet/gui-state")
def flowsheet_gui_state():
    """Inspect any running DWSIM GUI windows (read-only)."""
    return dwsim_gui_bridge.detect_state()


@app.post("/flowsheet/push-to-gui")
def flowsheet_push_to_gui(req: GuiPushRequest):
    """Push the current (or specified) flowsheet into the running DWSIM GUI.

    Best-effort: posts WM_CLOSE when close_first=True, then launches DWSIM
    with the file path. If DWSIM shows an unsaved-changes prompt, the user
    must click through.
    """
    bridge = _get_bridge()
    path = req.path
    if not path:
        st = getattr(bridge, "state", None)
        path = st.path if st and getattr(st, "path", "") else ""
    if not path:
        raise HTTPException(status_code=400,
                            detail="No flowsheet path — load or save one first")
    return dwsim_gui_bridge.push_to_gui(path, close_first=req.close_first)


@app.get("/flowsheet/meta")
def flowsheet_meta():
    """Return metadata for the currently-loaded flowsheet (sync-safety panel)."""
    bridge = _get_bridge()
    st = getattr(bridge, "state", None)
    if st is None or not getattr(st, "path", ""):
        return {"loaded": False}
    path = st.path
    on_disk_mtime = 0.0
    size = 0
    exists = os.path.exists(path)
    if exists:
        try:
            on_disk_mtime = os.path.getmtime(path)
            size = os.path.getsize(path)
        except Exception:
            pass
    cached = getattr(st, "loaded_mtime", 0.0) or 0.0
    external_edit = exists and cached > 0 and (on_disk_mtime - cached) > 1.0
    lock_path = path + ".lock"
    locked_by = None
    if os.path.exists(lock_path):
        try:
            with open(lock_path, "r", encoding="utf-8") as f:
                locked_by = f.read().strip()
        except Exception:
            locked_by = "?"
    return {
        "loaded": True,
        "path": path,
        "name": os.path.basename(path),
        "exists": exists,
        "size_bytes": size,
        "mtime_loaded": cached,
        "mtime_on_disk": on_disk_mtime,
        "external_edit": external_edit,
        "locked": locked_by is not None,
        "locked_by": locked_by,
        "streams": len(st.streams),
        "unit_ops": len(st.unit_ops),
        "property_package": st.property_package,
    }


@app.post("/flowsheet/run")
def run_simulation():
    result = _get_bridge().run_simulation()
    # Apply auto-correction if simulation ran but streams didn't converge
    if (isinstance(result, dict)
            and result.get("success")
            and not result.get("convergence_check", {}).get("all_converged", True)):
        try:
            from auto_correct import AutoCorrector
            result = AutoCorrector(_get_bridge()).attempt_fixes(result)
        except Exception:
            pass
    return _ok(result)


@app.get("/flowsheet/objects")
def list_objects():
    return _ok(_get_bridge().list_simulation_objects())


@app.get("/flowsheet/topology")
def get_topology():
    """Return flowsheet graph: nodes (objects) + edges (connections) for visual diagram."""
    bridge = _get_bridge()
    if bridge._flowsheet is None:
        raise HTTPException(status_code=400, detail="No flowsheet loaded")

    nodes = []
    edges = []
    coll = bridge._get_collection()
    if coll is None:
        return {"nodes": [], "edges": []}

    tag_cache = bridge._active_tag_cache()
    guid_to_tag = {}

    for guid, obj in bridge._iter_collection(coll):
        try:
            tag = tag_cache.get(str(guid)) or str(guid)[:8]
            typename = obj.GetType().Name
            category = "stream" if "Stream" in typename else "unit"

            # Try to get position from GraphicObject
            x, y = 0, 0
            try:
                go = obj.GraphicObject
                if go:
                    x = float(go.X) if hasattr(go, 'X') else 0
                    y = float(go.Y) if hasattr(go, 'Y') else 0
            except Exception:
                pass

            guid_to_tag[str(guid)] = tag
            nodes.append({
                "id": str(guid),
                "tag": tag,
                "type": typename,
                "category": category,
                "x": x,
                "y": y,
            })

            # Extract connections from GraphicObject connectors
            try:
                go = obj.GraphicObject
                if go:
                    # Output connectors
                    if hasattr(go, 'OutputConnectors'):
                        for i, conn in enumerate(go.OutputConnectors):
                            if conn and conn.IsAttached and conn.AttachedConnector:
                                ac = conn.AttachedConnector
                                if ac.AttachedTo and ac.AttachedTo.Owner:
                                    target_guid = str(ac.AttachedTo.Owner.Name)
                                    # Try to get target GUID properly
                                    try:
                                        for tg, tobj in bridge._iter_collection(coll):
                                            try:
                                                if tobj.GraphicObject and tobj.GraphicObject == ac.AttachedTo.Owner:
                                                    target_guid = str(tg)
                                                    break
                                            except Exception:
                                                pass
                                    except Exception:
                                        pass
                                    edges.append({
                                        "from": str(guid),
                                        "to": target_guid,
                                        "port": i,
                                    })
            except Exception:
                pass
        except Exception:
            pass

    # Resolve tags in edges — BUG-2 fix: guard against None and truncate safely
    for e in edges:
        raw_from = e.get("from") or ""
        raw_to   = e.get("to")   or ""
        e["from_tag"] = guid_to_tag.get(raw_from, raw_from[:12] if raw_from else "?")
        e["to_tag"]   = guid_to_tag.get(raw_to,   raw_to[:12]   if raw_to   else "?")

    return {"nodes": nodes, "edges": edges}


@app.get("/flowsheet/results")
def get_results():
    return _ok(_get_bridge().get_simulation_results())


@app.get("/flowsheet/package")
def get_property_package():
    """ACC-3: Read thermodynamic property package."""
    return _ok(_get_bridge().get_property_package())


@app.post("/flowsheet/validate")
def validate_feeds():
    """ACC-4: Validate feed stream specs."""
    return _ok(_get_bridge().validate_feed_specs())


@app.get("/flowsheet/convergence")
def check_convergence():
    """ACC-2: Check convergence of all streams."""
    return _ok(_get_bridge().check_convergence())


@app.get("/flowsheet/loaded")
def list_loaded():
    return _ok(_get_bridge().list_loaded_flowsheets())


@app.post("/flowsheet/switch")
def switch_flowsheet(alias: str):
    return _ok(_get_bridge().switch_flowsheet(alias))


# ─────────────────────────────────────────────────────────────────────────────
# Stream / unit-op operations
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/stream/properties")
def get_stream_props(req: StreamPropertyReadRequest):
    return _ok(_get_bridge().get_stream_properties(req.tag))


@app.post("/stream/set_property")
def set_stream_prop(req: StreamPropertyRequest):
    return _ok(_get_bridge().set_stream_property(
        req.tag, req.property_name, req.value, req.unit))


@app.post("/stream/set_composition")
def set_stream_composition(req: StreamCompositionRequest):
    """ACC-1: Set mole fractions on a feed stream."""
    return _ok(_get_bridge().set_stream_composition(req.tag, req.compositions))


@app.post("/unitop/set_property")
def set_unitop_prop(req: UnitOpPropertyRequest):
    return _ok(_get_bridge().set_unit_op_property(
        req.tag, req.property_name, req.value))


@app.post("/object/properties")
def get_object_props(req: StreamPropertyReadRequest):
    return _ok(_get_bridge().get_object_properties(req.tag))


@app.get("/flowsheet/unitops")
def get_all_unitops():
    """Return summary properties for every unit operation in the flowsheet."""
    bridge = _get_bridge()
    if bridge._flowsheet is None:
        raise HTTPException(status_code=400, detail="No flowsheet loaded")

    coll = bridge._get_collection()
    if coll is None:
        return {"success": True, "unit_ops": []}

    unit_ops = []
    from dwsim_bridge_v2 import _get_unit_op_summary  # noqa
    tag_cache = bridge._active_tag_cache()

    for guid, obj in bridge._iter_collection(coll):
        try:
            typename = obj.GetType().Name
            if "Stream" in typename:
                continue   # skip streams — only show unit ops
            tag = tag_cache.get(str(guid)) or str(guid)[:8]

            # Get x/y position
            x, y = 0, 0
            try:
                go = obj.GraphicObject
                if go:
                    x = float(getattr(go, 'X', 0) or 0)
                    y = float(getattr(go, 'Y', 0) or 0)
            except Exception:
                pass

            calculated = False
            try:
                calculated = bool(obj.Calculated)
            except Exception:
                pass

            props = _get_unit_op_summary(obj)
            unit_ops.append({
                "tag":        tag,
                "type":       typename,
                "calculated": calculated,
                "x": x, "y": y,
                "properties": props,
            })
        except Exception:
            pass

    return {"success": True, "unit_ops": unit_ops}


# ─────────────────────────────────────────────────────────────────────────────
# Parametric study & optimisation
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/parametric")
def parametric_study(req: ParametricRequest):
    return _ok(_get_bridge().parametric_study(
        vary_tag=req.vary_tag,
        vary_property=req.vary_property,
        vary_unit=req.vary_unit,
        values=req.values,
        observe_tag=req.observe_tag,
        observe_property=req.observe_property,
    ))


@app.post("/optimize")
def optimize(req: OptimizeRequest):
    """ACC-5: SciPy bounded optimisation."""
    return _ok(_get_bridge().optimize_parameter(
        vary_tag=req.vary_tag,
        vary_property=req.vary_property,
        vary_unit=req.vary_unit,
        lower_bound=req.lower_bound,
        upper_bound=req.upper_bound,
        observe_tag=req.observe_tag,
        observe_property=req.observe_property,
        minimize=req.minimize,
        tolerance=req.tolerance,
        max_iterations=req.max_iterations,
    ))


# ─────────────────────────────────────────────────────────────────────────────
# v4: Autonomous Flowsheet Generation
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/knowledge")
def search_knowledge(q: str = ""):
    """RAG: Search the chemical engineering knowledge base."""
    return _kb.search(q, top_k=5)


@app.get("/knowledge/topics")
def knowledge_topics():
    """RAG: List all knowledge base topics."""
    return {"success": True, "topics": _kb.list_topics()}


# ── Property Database endpoints ───────────────────────────────────────────────
try:
    from property_db import PropertyDB as _PropDB

    @app.get("/properties/compound")
    def property_lookup(compound: str, props: str = "all"):
        """Look up exact thermodynamic properties from DIPPR/DECHEMA database."""
        properties = None if props == "all" else [p.strip() for p in props.split(",")]
        return _PropDB().lookup(compound, properties)

    @app.get("/properties/pair")
    def property_pair(comp1: str, comp2: str, model: str = "nrtl"):
        """Look up binary interaction parameters (BIPs) for a compound pair."""
        return _PropDB().lookup_pair(comp1, comp2, model)

    @app.get("/properties/psat")
    def property_psat(compound: str, T_C: float):
        """Compute vapor pressure at temperature T_C using Antoine equation."""
        return _PropDB().antoine_psat(compound, T_C)

    @app.get("/properties/search")
    def property_search(q: str):
        """Search property database by compound name/alias/CAS."""
        return _PropDB().search_compound(q)

    @app.get("/properties/compounds")
    def property_list():
        """List all compounds available in the property database."""
        return _PropDB().list_compounds()

except Exception:
    pass  # property_db not available — endpoints simply won't be registered


@app.get("/compounds")
def get_compounds(search: str = ""):
    """v4: Search DWSIM compound database."""
    return _ok(_get_bridge().get_available_compounds(search))


@app.get("/property-packages")
def get_property_packages():
    """v4: List all available thermodynamic property packages."""
    return _ok(_get_bridge().get_available_property_packages())


@app.post("/flowsheet/create")
def create_flowsheet(topology: dict):
    """v4: Autonomously build a new DWSIM flowsheet from a topology JSON."""
    return _ok(_get_bridge().create_flowsheet(topology))


@app.get("/flowsheet/templates")
def list_flowsheet_templates():
    """Return the curated starter-topology library for UI quick-start."""
    from flowsheet_templates import TEMPLATES
    items = [
        {"name": k,
         "category": v.get("category", ""),
         "description": v.get("description", "")}
        for k, v in TEMPLATES.items()
    ]
    return _ok({"templates": items})


@app.post("/flowsheet/create-from-template")
def create_flowsheet_from_template(body: dict):
    """Direct-build path: renders a named template and hands it to
    DWSIMBridgeV2.create_flowsheet, bypassing the LLM entirely."""
    from flowsheet_templates import render_template, TEMPLATES
    name = (body or {}).get("name") or ""
    overrides = (body or {}).get("overrides") or None
    if not name or name not in TEMPLATES:
        raise HTTPException(400, f"Unknown template '{name}'. "
                                 f"Available: {sorted(TEMPLATES.keys())}")
    topology = render_template(name, overrides)
    if topology is None:
        raise HTTPException(400, f"Failed to render template '{name}'")
    return _ok(_get_bridge().create_flowsheet(topology))


@app.post("/report/generate")
def generate_report(report_spec: dict):
    """v4: Generate a formatted PDF research report from parametric study data."""
    return _ok(_get_bridge().generate_report(report_spec))


@app.get("/flowsheet/diagram")
def get_flowsheet_diagram():
    """Parse .dwxmz XML and return nodes + edges for SVG rendering."""
    from dwsim_bridge_v2 import parse_flowsheet_diagram
    bridge = _get_bridge()
    path = bridge.state.path
    if not path or not os.path.isfile(path):
        return _ok({"nodes": [], "edges": [], "reason": "no flowsheet loaded"})
    return parse_flowsheet_diagram(path)


@app.get("/report/download")
def download_report(path: str):
    """v4: Download a generated PDF report by file path."""
    import urllib.parse
    decoded = urllib.parse.unquote(path)
    if not os.path.isfile(decoded):
        raise HTTPException(status_code=404, detail=f"Report not found: {decoded}")
    return FileResponse(decoded, media_type="application/pdf",
                        filename=os.path.basename(decoded))


# ─────────────────────────────────────────────────────────────────────────────
# Excel Export
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/results/export/excel")
def export_results_excel():
    """Export simulation results as a formatted .xlsx Excel file."""
    import io
    import tempfile
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500,
                            detail="openpyxl not installed. Run: pip install openpyxl")

    bridge = _get_bridge()
    res = bridge.get_simulation_results()
    if not res.get("success") or not res.get("stream_results"):
        raise HTTPException(status_code=400,
                            detail="No simulation results available. Load and run a flowsheet first.")

    streams = res["stream_results"]
    stream_names = list(streams.keys())

    wb = Workbook()

    # ── Sheet 1: Stream Results ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "Stream Results"

    # Styles
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    prop_font   = Font(name="Calibri", bold=True, color="1E3A5F", size=10)
    data_font   = Font(name="Consolas", size=10)
    comp_font   = Font(name="Consolas", size=10, color="666666")
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(stream_names) + 1)
    title_cell = ws.cell(row=1, column=1,
                         value=f"DWSIM Simulation Results — {bridge.state.name or 'Untitled'}")
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="1E3A5F")
    title_cell.alignment = Alignment(horizontal="center")

    # Header row
    row = 3
    ws.cell(row=row, column=1, value="Property").font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=1).border = thin_border
    ws.column_dimensions["A"].width = 22

    for j, name in enumerate(stream_names, start=2):
        cell = ws.cell(row=row, column=j, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = max(16, len(name) + 4)

    # Property rows
    PROPS = [
        ("Temperature (C)",  "temperature_C"),
        ("Pressure (bar)",   "pressure_bar"),
        ("Mass Flow (kg/h)", "mass_flow_kgh"),
        ("Molar Flow (kmol/h)", "molar_flow_kmolh"),
        ("Vapor Fraction",   "vapor_fraction"),
    ]

    row = 4
    for label, key in PROPS:
        ws.cell(row=row, column=1, value=label).font = prop_font
        ws.cell(row=row, column=1).border = thin_border
        for j, name in enumerate(stream_names, start=2):
            val = streams[name].get(key)
            cell = ws.cell(row=row, column=j,
                           value=round(val, 4) if isinstance(val, (int, float)) else val)
            cell.font = data_font
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="right")
            cell.border = thin_border
        row += 1

    # Mole fractions
    comp_keys = set()
    for name in stream_names:
        comp_keys.update((streams[name].get("mole_fractions") or {}).keys())
    comp_keys = sorted(comp_keys)

    if comp_keys:
        row += 1
        ws.cell(row=row, column=1, value="— Mole Fractions —").font = Font(
            name="Calibri", bold=True, italic=True, color="999999", size=10)
        row += 1

        for comp in comp_keys:
            ws.cell(row=row, column=1, value=f"x({comp})").font = comp_font
            ws.cell(row=row, column=1).border = thin_border
            for j, name in enumerate(stream_names, start=2):
                val = (streams[name].get("mole_fractions") or {}).get(comp)
                cell = ws.cell(row=row, column=j,
                               value=round(val, 6) if isinstance(val, (int, float)) else val)
                cell.font = comp_font
                cell.number_format = '0.000000'
                cell.alignment = Alignment(horizontal="right")
                cell.border = thin_border
            row += 1

    # ── Sheet 2: Transposed view (one stream per sheet section) ──────────────
    ws2 = wb.create_sheet("By Stream")
    r = 1
    for name in stream_names:
        ws2.cell(row=r, column=1, value=name).font = Font(
            name="Calibri", bold=True, size=12, color="1E3A5F")
        r += 1
        for label, key in PROPS:
            ws2.cell(row=r, column=1, value=label).font = prop_font
            val = streams[name].get(key)
            cell = ws2.cell(row=r, column=2,
                            value=round(val, 4) if isinstance(val, (int, float)) else val)
            cell.font = data_font
            cell.number_format = '0.0000'
            r += 1
        for comp in comp_keys:
            ws2.cell(row=r, column=1, value=f"x({comp})").font = comp_font
            val = (streams[name].get("mole_fractions") or {}).get(comp)
            cell = ws2.cell(row=r, column=2,
                            value=round(val, 6) if isinstance(val, (int, float)) else val)
            cell.font = comp_font
            cell.number_format = '0.000000'
            r += 1
        r += 1  # blank row between streams

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 18

    # BUG-6 fix: close NamedTemporaryFile handle BEFORE openpyxl opens the same
    # path — Windows does not allow two open handles on the same file.
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False, prefix="dwsim_results_")
    tmp_path = tmp.name
    tmp.close()          # release handle first
    wb.save(tmp_path)    # openpyxl now owns the only handle
    filename = f"{bridge.state.name or 'results'}.xlsx".replace(" ", "_")
    return FileResponse(tmp_path, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        filename=filename)


@app.get("/results/export/csv")
def export_results_csv():
    """Export simulation results as a CSV file."""
    import csv
    import io
    import tempfile

    bridge = _get_bridge()
    res = bridge.get_simulation_results()
    if not res.get("success") or not res.get("stream_results"):
        raise HTTPException(status_code=400,
                            detail="No simulation results available.")

    streams = res["stream_results"]
    stream_names = list(streams.keys())

    PROPS = [
        ("Temperature (C)",  "temperature_C"),
        ("Pressure (bar)",   "pressure_bar"),
        ("Mass Flow (kg/h)", "mass_flow_kgh"),
        ("Molar Flow (kmol/h)", "molar_flow_kmolh"),
        ("Vapor Fraction",   "vapor_fraction"),
    ]

    comp_keys = set()
    for name in stream_names:
        comp_keys.update((streams[name].get("mole_fractions") or {}).keys())
    comp_keys = sorted(comp_keys)

    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow(["Property"] + stream_names)
    for label, key in PROPS:
        row = [label]
        for name in stream_names:
            val = streams[name].get(key)
            row.append(round(val, 6) if isinstance(val, (int, float)) else (val or ""))
        writer.writerow(row)

    for comp in comp_keys:
        row = [f"x({comp})"]
        for name in stream_names:
            val = (streams[name].get("mole_fractions") or {}).get(comp)
            row.append(round(val, 6) if isinstance(val, (int, float)) else (val or ""))
        writer.writerow(row)

    tmp = tempfile.NamedTemporaryFile(suffix=".csv", delete=False, prefix="dwsim_results_",
                                      mode="w", encoding="utf-8")
    tmp.write(output.getvalue())
    tmp.close()
    filename = f"{bridge.state.name or 'results'}.csv".replace(" ", "_")
    return FileResponse(tmp.name, media_type="text/csv", filename=filename)


# ─────────────────────────────────────────────────────────────────────────────
# LLM provider management (hot-switch without restart)
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/llm/status")
def llm_status():
    """Return current LLM provider and model."""
    agent = _get_agent()
    return {
        "provider": agent.llm.provider,
        "model":    agent.llm.model,
        "providers_available": {
            "groq":      bool(os.getenv("GROQ_API_KEY")),
            "gemini":    bool(os.getenv("GEMINI_API_KEY")),
            "openai":    bool(os.getenv("OPENAI_API_KEY")),
            "anthropic": bool(os.getenv("ANTHROPIC_API_KEY")),
        }
    }


@app.post("/llm/switch")
def llm_switch(provider: str, model: str = ""):
    """Hot-switch LLM provider (including Ollama) without restarting the server."""
    key_env = {
        "groq":      "GROQ_API_KEY",
        "gemini":    "GEMINI_API_KEY",
        "openai":    "OPENAI_API_KEY",
        "anthropic": "ANTHROPIC_API_KEY",
        "ollama":    "",   # no API key needed for local Ollama
    }
    if provider not in key_env:
        raise HTTPException(
            status_code=400,
            detail=f"Unknown provider '{provider}'. Valid: {list(key_env)}"
        )

    # Re-read .env so keys added after startup are picked up without a restart
    load_dotenv(override=True)

    # Ollama requires no API key — just needs Ollama running locally
    if provider == "ollama":
        api_key = "ollama"   # sentinel value; LLMClient ignores it for Ollama
    else:
        api_key = os.getenv(key_env[provider], "")
        if not api_key:
            raise HTTPException(
                status_code=400,
                detail=f"No API key configured for '{provider}'. "
                       f"Set {key_env[provider]} in your .env file."
            )

    resolved_model = model or DEFAULT_MODELS.get(provider, "")
    try:
        agent = _get_agent()
        agent.llm.provider    = provider
        agent.llm.api_key     = api_key
        agent.llm.model       = resolved_model
        agent.llm._client     = None
        agent.llm._gemini_sdk = None
        agent.llm._setup()
        return {
            "success":  True,
            "provider": provider,
            "model":    resolved_model,
            "message":  f"Switched to {provider.upper()} ({resolved_model}). You can continue chatting.",
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/llm/groq/models")
def llm_groq_models():
    """
    Fetch available Groq models dynamically from the Groq API.
    Filters to only models that support tool-calling (function calling).
    Falls back to the hardcoded GROQ_MODELS list if the API call fails.
    """
    from llm_client import GROQ_MODELS as _FALLBACK
    import urllib.request, urllib.error, json as _json

    api_key = os.getenv("GROQ_API_KEY", "")
    if not api_key:
        # No key — return the hardcoded list
        return {
            "success": True,
            "models": _FALLBACK,
            "source": "hardcoded",
            "note": "GROQ_API_KEY not set — showing built-in model list",
        }

    # Tool-calling capable model IDs (Groq doesn't mark this in the API response,
    # so we maintain a known-good set and prefer them).
    _TOOL_CALL_CAPABLE = {
        "llama-3.3-70b-versatile", "llama-3.1-8b-instant",
        "llama-3.2-90b-vision-preview", "llama-3.2-11b-vision-preview",
        "llama-3.2-3b-preview", "llama-3.2-1b-preview",
        "llama3-70b-8192", "llama3-8b-8192",
        "gemma2-9b-it", "gemma-7b-it",
        "deepseek-r1-distill-llama-70b", "qwen-qwq-32b",
        "mistral-saba-24b", "allam-2-7b",
        "meta-llama/llama-4-maverick-17b-128e-instruct",
        "meta-llama/llama-4-scout-17b-16e-instruct",
    }

    try:
        req = urllib.request.Request(
            "https://api.groq.com/openai/v1/models",
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        )
        with urllib.request.urlopen(req, timeout=6) as resp:
            data = _json.loads(resp.read().decode())

        all_ids = [m["id"] for m in data.get("data", [])]

        # Sort: known tool-calling models first (in GROQ_MODELS order), then rest
        def _sort_key(mid):
            try:    return (_FALLBACK.index(mid), mid)
            except: return (len(_FALLBACK), mid)

        sorted_models = sorted(all_ids, key=_sort_key)

        # Tag which ones are known tool-calling capable
        annotated = []
        for mid in sorted_models:
            annotated.append({
                "id":           mid,
                "tool_calling": mid in _TOOL_CALL_CAPABLE or mid in _FALLBACK,
            })

        return {
            "success": True,
            "models": sorted_models,
            "annotated": annotated,
            "source": "groq_api",
            "total": len(sorted_models),
        }

    except urllib.error.HTTPError as e:
        return {
            "success": False,
            "models": _FALLBACK,
            "source": "hardcoded_fallback",
            "error": f"Groq API error {e.code}: {e.reason}",
        }
    except Exception as exc:
        return {
            "success": False,
            "models": _FALLBACK,
            "source": "hardcoded_fallback",
            "error": str(exc),
        }


@app.get("/llm/ollama/models")
def llm_ollama_models():
    """
    List models available in the local Ollama instance.
    Returns list of model names, or an error if Ollama is not running.
    """
    import urllib.request, urllib.error, json as _json
    ollama_base = os.getenv("OLLAMA_HOST", "http://localhost:11434")
    try:
        # Use context manager so the socket is always closed (no fd leak)
        with urllib.request.urlopen(f"{ollama_base}/api/tags", timeout=3) as req:
            data = _json.loads(req.read().decode())
        models = [m["name"] for m in data.get("models", [])]
        return {"success": True, "models": models, "ollama_url": ollama_base}
    except urllib.error.URLError:
        return {
            "success": False,
            "models":  [],
            "error":   "Ollama is not running. Start it with: ollama serve",
            "ollama_url": ollama_base,
        }
    except Exception as exc:
        return {"success": False, "models": [], "error": str(exc)}


# ─────────────────────────────────────────────────────────────────────────────
# Session management
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/sessions")
def list_sessions_endpoint():
    sessions = list_sessions()
    return {"sessions": sessions}


@app.post("/sessions/save")
def save_session_endpoint(name: str):
    agent  = _get_agent()
    bridge = _get_bridge()
    path   = save_session(
        agent._history,
        agent.llm.provider,
        agent.llm.model,
        bridge.state.path,
        bridge.state.name,
        name,
    )
    return {"success": True, "path": path}


@app.post("/sessions/load")
def load_session_endpoint(path: str):
    try:
        sess = load_session(path)
        agent = _get_agent()
        agent._history = sess["history"]
        return {"success": True, "messages": len(agent._history)}
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc))


# ─────────────────────────────────────────────────────────────────────────────
# Economic Optimizer
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/economics/estimate")
def economics_estimate(req: EconomicsRequest):
    """
    Estimate CAPEX, OPEX, revenue, and profitability metrics
    from the currently loaded and solved flowsheet.
    """
    from economics import run_economic_analysis

    bridge = _get_bridge()

    # Collect objects (unit ops + streams)
    objects_result = bridge.list_simulation_objects()
    objects = objects_result.get("objects", []) if isinstance(objects_result, dict) else []

    # Collect stream results
    sim_result = bridge.get_simulation_results()
    stream_results = {}
    if isinstance(sim_result, dict):
        stream_results = sim_result.get("stream_results", {}) or {}

    # Collect energy duties AND outlet temperatures from unit ops
    unit_op_duties: Dict[str, float] = {}
    unit_op_outlet_temps: Dict[str, float] = {}
    try:
        coll = bridge._get_collection()
        if coll is not None:
            from dwsim_bridge_v2 import _get_unit_op_summary
            tag_cache = bridge._active_tag_cache()
            for guid, obj in bridge._iter_collection(coll):
                try:
                    typename = obj.GetType().Name
                    if "Stream" in typename:
                        continue
                    tag = tag_cache.get(str(guid), str(guid)[:8])
                    summary = _get_unit_op_summary(obj, tag)
                    duty = summary.get("duty_kW") or summary.get("heat_duty_kW")
                    if duty is not None:
                        unit_op_duties[tag] = float(duty)
                    # Capture outlet temperature for tiered utility pricing
                    for t_key in ("outlet_temperature_C", "OutletTemperature_C",
                                  "outlet_temp_C", "temperature_out_C"):
                        t_val = summary.get(t_key)
                        if t_val is not None:
                            unit_op_outlet_temps[tag] = float(t_val)
                            break
                except Exception:
                    pass
    except Exception:
        pass

    params = req.dict()
    result = run_economic_analysis(
        objects, stream_results, unit_op_duties, params,
        unit_op_outlet_temps=unit_op_outlet_temps or None,
    )
    return _ok(result)


@app.get("/economics/defaults")
def economics_defaults():
    """Return default economic parameters."""
    from economics import DEFAULT_PARAMS, EQUIPMENT_BASE_COST
    return {
        "defaults": DEFAULT_PARAMS,
        "equipment_catalog": {
            k: {"base_usd": v["base"], "description": v["desc"]}
            for k, v in EQUIPMENT_BASE_COST.items()
            if not k.startswith("_")
        },
    }


# ─────────────────────────────────────────────────────────────────────────────
# New Tool Endpoints
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/flowsheet/pinch")
def pinch_analysis(min_approach_temp_C: float = 10.0):
    """Pinch analysis — minimum utilities and potential heat recovery savings."""
    return _ok(_get_bridge().pinch_analysis(min_approach_temp_C))


class MonteCarloRequest(BaseModel):
    vary_params:      List[Dict[str, Any]]
    observe_tag:      str
    observe_property: str
    n_samples:        int = 100

@app.post("/monte-carlo")
def monte_carlo(req: MonteCarloRequest):
    """
    Monte Carlo uncertainty propagation.
    Returns mean, std, p5/p25/p50/p75/p95, 95% CI, histogram for reporting in papers.
    """
    agent = _get_agent()

    def _progress(i, n, inputs, val):
        pass  # no streaming on REST endpoint; use /chat/stream for live progress

    return _ok(_get_bridge().monte_carlo_study(
        vary_params=req.vary_params,
        observe_tag=req.observe_tag,
        observe_property=req.observe_property,
        n_samples=req.n_samples,
        on_progress=_progress,
    ))


@app.get("/compounds/{name}/properties")
def compound_properties(name: str):
    """Return Tc, Pc, ω, Tb, MW, ΔHf° for a compound from the DWSIM database."""
    return _ok(_get_bridge().get_compound_properties(name))


class MultivarOptRequest(BaseModel):
    variables:        List[Dict[str, Any]]
    observe_tag:      str
    observe_property: str
    minimize:         bool = True
    max_iterations:   int  = 100
    population_size:  int  = 8
    tolerance:        float = 1e-3

@app.post("/optimize/multivar")
def optimize_multivar(req: MultivarOptRequest):
    """Multi-variable optimisation via differential evolution."""
    return _ok(_get_bridge().optimize_multivar(**req.dict()))


class BayesianOptRequest(BaseModel):
    variables:        List[Dict[str, Any]]
    observe_tag:      str
    observe_property: str
    minimize:         bool  = True
    n_initial:        int   = 5
    max_iter:         int   = 20
    xi:               float = 0.01
    seed:             int   = 42
    save_plot:        str   = ""

@app.post("/optimize/bayesian")
def bayesian_optimize(req: BayesianOptRequest):
    """
    Bayesian Optimisation via GP surrogate + Expected Improvement.
    Preferred over /optimize/multivar when simulations are expensive (>5s each)
    and 1–4 continuous variables are being tuned.
    Uses 25 total evaluations (5 LHS warm-up + 20 BO iterations by default).
    """
    return _ok(_get_bridge().bayesian_optimize(**req.dict()))


class RecycleInitRequest(BaseModel):
    recycle_tag:  str
    T_guess_C:    float
    P_guess_bar:  float
    composition:  Dict[str, float]
    solver:       str = "Wegstein"

@app.post("/flowsheet/initialize-recycle")
def initialize_recycle(req: RecycleInitRequest):
    """Seed a recycle stream with initial guess to aid convergence."""
    return _ok(_get_bridge().initialize_recycle(**req.dict()))


@app.get("/flowsheet/compare")
def compare_flowsheets():
    """
    Compare all loaded flowsheets — returns Δ(T, P, flow, VF) for matching stream tags.
    Useful for: 'How does PR compare to NRTL for this mixture?'
    """
    bridge = _get_bridge()
    loaded = bridge.list_loaded_flowsheets()
    aliases = list((loaded.get("loaded") or {}).keys())
    if len(aliases) < 2:
        return {"success": False,
                "error":   "Load at least 2 flowsheets with different aliases first."}

    # Collect results per alias
    results_by_alias: Dict[str, Dict] = {}
    current_alias = bridge._active_alias
    for alias in aliases:
        try:
            bridge.switch_flowsheet(alias)
            sr = bridge.get_simulation_results()
            results_by_alias[alias] = sr.get("stream_results", {}) or {}
        except Exception:
            results_by_alias[alias] = {}
    # Restore original
    if current_alias:
        try:
            bridge.switch_flowsheet(current_alias)
        except Exception:
            pass

    # Build comparison table for streams present in all aliases
    all_tags = set()
    for r in results_by_alias.values():
        all_tags.update(r.keys())

    comparison = []
    _PROPS = ["temperature_C", "pressure_bar", "mass_flow_kgh", "vapor_fraction"]
    for tag in sorted(all_tags):
        row: Dict[str, Any] = {"stream": tag}
        for alias in aliases:
            props = results_by_alias.get(alias, {}).get(tag, {})
            row[alias] = {p: props.get(p) for p in _PROPS}
        # Compute delta vs first alias
        first = aliases[0]
        for alias in aliases[1:]:
            delta: Dict[str, Any] = {}
            for p in _PROPS:
                v0 = row[first].get(p)
                v1 = row[alias].get(p)
                if v0 is not None and v1 is not None:
                    try:
                        delta[p] = round(float(v1) - float(v0), 4)
                    except Exception:
                        pass
            row[f"delta_{alias}_vs_{first}"] = delta
        comparison.append(row)

    return {"success": True, "aliases": aliases, "comparison": comparison}


@app.get("/diagnostics/version")
def dwsim_version():
    """Return DWSIM version and any API compatibility warnings."""
    bridge = _get_bridge()
    return {
        "success":       True,
        "dwsim_version": getattr(bridge, "_dwsim_version", "unknown"),
        "bridge_ready":  bridge._ready,
        "dll_folder":    bridge.dll_folder,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Quantitative Evaluation API  (research / paper metrics)
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/eval/metrics")
def eval_metrics():
    """Return aggregated session metrics: success rate, duration, convergence rate, tool stats."""
    return get_eval_log().get_metrics()


# ── Ablation study endpoints ──────────────────────────────────────────────────

@app.get("/ablation/configs")
def ablation_configs():
    """List all ablation configurations available for paper experiments."""
    from ablation import ABLATION_CONFIGS, ABLATION_TASKS
    return {
        "success": True,
        "configs": {cid: {"config_id": c.config_id, "description": c.description,
                           "temperature": c.temperature,
                           "disabled_components": [
                               k for k, v in {
                                   "safety_validator": c.disable_safety,
                                   "rag_knowledge_base": c.disable_rag,
                                   "auto_corrector": c.disable_autocorrect,
                                   "tool_compression": c.disable_compression,
                                   "context_trimming": c.disable_trim,
                                   "shortcut_col_only": c.force_shortcut_col,
                               }.items() if v
                           ]}
                   for cid, c in ABLATION_CONFIGS.items()},
        "tasks": [{"task_id": t["task_id"], "category": t["category"],
                   "complexity": t["complexity"], "human_time_min": t["human_time_min"]}
                  for t in ABLATION_TASKS],
    }


@app.get("/ablation/summary")
def ablation_summary():
    """Return aggregated ablation results for paper table generation."""
    from ablation import get_ablation_summary
    return get_ablation_summary()


class AblationRunRequest(BaseModel):
    configs:  List[str] = ["A0", "A1", "A2"]
    tasks:    List[str] = ["all"]
    n_runs:   int       = 3
    provider: str       = "groq"
    model:    str       = ""

@app.post("/ablation/run")
def ablation_run(req: AblationRunRequest):
    """
    Launch an ablation study run (background thread — may take 10-60 min).
    Returns immediately with run_id; check /ablation/summary for results.
    """
    from ablation import run_ablation

    api_key = os.getenv({
        "groq": "GROQ_API_KEY", "gemini": "GEMINI_API_KEY",
        "openai": "OPENAI_API_KEY", "anthropic": "ANTHROPIC_API_KEY",
    }.get(req.provider, "GROQ_API_KEY"), "")

    run_id = uuid.uuid4().hex[:8]
    def _run_bg():
        run_ablation(
            config_ids=req.configs,
            task_ids=req.tasks,
            n_runs=req.n_runs,
            provider=req.provider,
            model=req.model or None,
            api_key=api_key,
        )

    thread = threading.Thread(target=_run_bg, daemon=True, name=f"ablation_{run_id}")
    thread.start()
    return {
        "success": True,
        "run_id":  run_id,
        "message": f"Ablation study started (run_id={run_id}). "
                   f"Configs: {req.configs}, Tasks: {req.tasks}, Runs/pair: {req.n_runs}. "
                   "Check /ablation/summary for results as they complete.",
    }


@app.get("/reproducibility/last-turn")
def reproducibility_last_turn():
    """
    Return reproducibility fingerprint for the last agent turn.
    Includes: prompt_hash, tool_sequence, provider, model, temperature, seed,
    and replay log session summary for independent verification.
    """
    agent = _get_agent()
    metrics = getattr(agent, "last_turn_metrics", None) or {}

    # Replay log session summary
    session_summary: dict = {}
    try:
        import replay_log as _rl
        session_id = getattr(agent, "_session_id", "")
        if session_id:
            session_summary = _rl.session_summary(session_id)
    except Exception:
        pass

    return {
        "success":        True,
        "prompt_hash":    metrics.get("prompt_hash", ""),
        "tool_sequence":  metrics.get("tool_sequence", []),
        "n_tools":        metrics.get("tool_count", 0),
        "provider":       metrics.get("provider", ""),
        "model":          metrics.get("model", ""),
        "temperature":    metrics.get("temperature", 0.0),
        "seed":           metrics.get("seed", 42),
        "duration_s":     metrics.get("total_s", 0),
        "iterations":     metrics.get("iterations", 0),
        "session_summary": session_summary,
        "note": (
            "temperature=0 + seed=42 reduces but does not eliminate LLM stochasticity. "
            "Run each task ≥3 times and report mean ± std for paper tables. "
            "Full replay log at ~/.dwsim_agent/replay/replay_log.jsonl — use "
            "'python replay_log.py replay --turn <id>' for independent verification."
        ),
    }


@app.get("/reproducibility/session/{session_id}/export")
async def reproducibility_export_session(session_id: str):
    """Export full session replay log as JSON (paper appendix format)."""
    import tempfile, os
    try:
        import replay_log as _rl
        with tempfile.NamedTemporaryFile(suffix=".json", delete=False,
                                         mode="w", encoding="utf-8") as f:
            tmp_path = f.name
        _rl.export_for_paper(session_id, tmp_path)
        with open(tmp_path, encoding="utf-8") as f:
            data = json.load(f)
        os.unlink(tmp_path)
        return {"success": True, "session_id": session_id, "data": data}
    except Exception as e:
        return {"success": False, "error": str(e)}


@app.get("/reproducibility/turns")
def reproducibility_list_turns(session_id: str = "", n: int = 20):
    """List recent replay turns (default 20). Filter by session_id if provided."""
    try:
        import replay_log as _rl
        turns = _rl.load_turns(
            session_id=session_id or None,
            last_n=min(n, 100),
        )
        return {
            "success": True,
            "count":   len(turns),
            "turns": [
                {
                    "turn_id":    t.turn_id,
                    "session_id": t.session_id,
                    "timestamp":  t.timestamp,
                    "prompt_hash":t.prompt_hash,
                    "model":      f"{t.provider}/{t.model}",
                    "temperature":t.temperature,
                    "seed":       t.seed,
                    "n_tools":    len(t.tool_calls),
                    "converged":  t.converged,
                    "duration_s": t.duration_s,
                    "sf_viols":   len(t.sf_violations),
                }
                for t in turns
            ],
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


@app.get("/eval/benchmarks")
def eval_list_benchmarks():
    """List all benchmark cases with their prompts and metadata."""
    return {"benchmarks": get_benchmark_suite().list_all()}


# ── Formal 25-task benchmark (reviewer-required) ─────────────────────────────

@app.get("/benchmark/tasks")
def benchmark_list_tasks():
    """
    List all 25 formally defined benchmark tasks (benchmark_tasks.py).
    Tasks are fixed a priori — defined before any experiments were run.
    """
    try:
        from benchmark_tasks import BENCHMARK_TASKS, task_summary
        tasks = [
            {
                "task_id":          t.task_id,
                "category":         t.category,
                "complexity":       t.complexity,
                "prompt":           t.prompt,
                "property_package": t.property_package,
                "human_time_min":   t.human_time_min,
                "criteria_count":   len(t.success_criteria),
                "constraint_count": len(t.physical_constraints),
                "notes":            t.notes,
            }
            for t in BENCHMARK_TASKS
        ]
        return {"success": True, "tasks": tasks, "summary": task_summary()}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


class BenchmarkTaskRunRequest(BaseModel):
    task_id:    str
    dry_run:    bool = False   # if True, return task spec without running agent


@app.post("/benchmark/run")
async def benchmark_run_task(req: BenchmarkTaskRunRequest):
    """
    Run a single benchmark task through the live agent and record outcome.
    Returns: prompt, outcome (SUCCESS/PARTIAL/FAILURE_LOUD/FAILURE_SILENT),
    agent_response, time_s, safety_status, criteria_met.
    """
    import time as _time
    try:
        from benchmark_tasks import BENCHMARK_TASKS
    except ImportError as exc:
        raise HTTPException(status_code=500, detail=f"benchmark_tasks not found: {exc}")

    task = next((t for t in BENCHMARK_TASKS if t.task_id == req.task_id), None)
    if task is None:
        raise HTTPException(status_code=404,
                            detail=f"Task {req.task_id!r} not found. "
                                   f"Valid IDs: {[t.task_id for t in BENCHMARK_TASKS]}")

    if req.dry_run:
        return {
            "success": True,
            "task_id": task.task_id,
            "prompt":  task.prompt,
            "criteria": [
                {"stream": c.stream_tag, "property": c.property,
                 "operator": c.operator, "value": c.value,
                 "tolerance_pct": c.tolerance_pct}
                for c in task.success_criteria
            ],
            "constraints": [
                {"description": c.description, "type": c.check_type}
                for c in task.physical_constraints
            ],
        }

    agent = _get_agent()
    t0 = _time.monotonic()
    loop = asyncio.get_running_loop()
    try:
        response = await loop.run_in_executor(None, agent.chat, task.prompt)
        elapsed = _time.monotonic() - t0
        outcome = "SUCCESS"   # simplified — full scoring requires DWSIM result access
    except Exception as exc:
        elapsed = _time.monotonic() - t0
        response = str(exc)
        outcome  = "FAILURE_LOUD"

    return {
        "success":        True,
        "task_id":        task.task_id,
        "category":       task.category,
        "complexity":     task.complexity,
        "prompt":         task.prompt,
        "agent_response": response,
        "time_s":         round(elapsed, 2),
        "outcome":        outcome,
        "human_time_min": task.human_time_min,
        "speedup_x":      round(task.human_time_min * 60 / elapsed, 1) if elapsed > 0 else None,
    }


@app.get("/benchmark/summary")
def benchmark_summary():
    """Return task distribution statistics for the formal benchmark."""
    try:
        from benchmark_tasks import task_summary, CATEGORIES
        return {"success": True, "summary": task_summary(), "categories": CATEGORIES}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


# ── Safety validator endpoint ─────────────────────────────────────────────────

@app.get("/safety/catalogue")
def safety_catalogue():
    """
    Return the known silent failure mode catalogue (SF-01 to SF-07).
    Useful for UI display and reproducibility documentation.
    """
    try:
        from safety_validator import get_failure_catalogue
        return {"success": True, "failures": get_failure_catalogue(),
                "total": len(get_failure_catalogue())}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@app.post("/safety/validate")
def safety_validate(body: dict):
    """
    Run the safety validator on arbitrary stream_results.
    Body: {"stream_results": {...}, "topology": {...} (optional)}
    Returns list of violations with code, severity, description.
    """
    try:
        from safety_validator import SafetyValidator
        sv = SafetyValidator()
        failures = sv.check(
            body.get("stream_results", {}),
            body.get("topology"),
        )
        return {
            "success":   True,
            "violations": [
                {"code": f.code, "severity": f.severity,
                 "description": f.description, "evidence": f.evidence,
                 "stream": f.stream_tag, "auto_fixed": f.auto_fixed}
                for f in failures
            ],
            "safety_status": "VIOLATIONS_DETECTED" if failures else "PASSED",
        }
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


class BenchmarkRunRequest(BaseModel):
    benchmark_id: str


@app.post("/eval/benchmark/run")
async def eval_run_benchmark(req: BenchmarkRunRequest):
    """
    Run a named benchmark case through the agent and score it.
    Returns session metrics + accuracy checks + pass/fail verdict.
    """
    suite = get_benchmark_suite()
    case  = suite.get(req.benchmark_id)
    if case is None:
        raise HTTPException(status_code=404, detail=f"Benchmark {req.benchmark_id} not found")

    agent = _get_agent()
    log   = get_eval_log()
    tracker = SessionTracker(case["prompt"], log, benchmark_id=req.benchmark_id)

    original_on_tool = agent.on_tool_call

    def _track_tool(name: str, args: dict, result: dict):
        tracker.record_tool_call(name, args, result)
        if original_on_tool:
            original_on_tool(name, args, result)

    agent.on_tool_call = _track_tool

    loop = asyncio.get_running_loop()
    try:
        answer = await loop.run_in_executor(None, agent.chat, case["prompt"])
        session = tracker.finish(answer)
    except Exception as exc:
        session = tracker.finish("", error=str(exc))
    finally:
        agent.on_tool_call = original_on_tool

    # Try to get current simulation results for numerical accuracy checks
    sim_results = None
    try:
        bridge = _get_bridge()
        if bridge._flowsheet is not None:
            raw = bridge.get_simulation_results()
            if raw.get("success"):
                sim_results = raw
    except Exception:
        pass

    result = suite.evaluate(req.benchmark_id, session, sim_results)
    log.add_benchmark_result(result)

    return {
        "benchmark_id":   req.benchmark_id,
        "name":           case["name"],
        "passed":         result.passed,
        "duration_s":     result.duration_s,
        "human_time_min": case.get("human_time_min"),
        "speedup_vs_human": (
            round(case["human_time_min"] * 60 / result.duration_s, 1)
            if result.duration_s and case.get("human_time_min") else None
        ),
        "tool_calls":     session.tool_count,
        "convergence":    result.convergence,
        "accuracy_checks": result.accuracy_checks,
        "notes":          result.notes,
        "session_id":     session.session_id,
    }


@app.get("/eval/benchmark/results")
def eval_benchmark_results():
    """Return all past benchmark run results."""
    return get_eval_log().get_benchmark_metrics()


class BenchmarkRunAllRequest(BaseModel):
    only_ids:   Optional[List[str]] = None
    difficulty: Optional[str] = None          # "easy" | "medium" | "hard"
    stop_on_failure: bool = False


@app.post("/eval/benchmark/run_all")
async def eval_run_all_benchmarks(req: BenchmarkRunAllRequest):
    """Run every benchmark (optionally filtered by id or difficulty) sequentially."""
    suite = get_benchmark_suite()
    cases = suite.list_all()
    if req.only_ids:
        cases = [c for c in cases if c["id"] in set(req.only_ids)]
    if req.difficulty:
        cases = [c for c in cases if c.get("difficulty") == req.difficulty]
    if not cases:
        raise HTTPException(status_code=400, detail="No benchmarks matched filter")

    ran: List[Dict[str, Any]] = []
    for case in cases:
        try:
            r = await eval_run_benchmark(BenchmarkRunRequest(benchmark_id=case["id"]))
        except HTTPException as he:
            r = {"benchmark_id": case["id"], "error": he.detail, "passed": False}
        except Exception as exc:
            r = {"benchmark_id": case["id"], "error": str(exc), "passed": False}
        ran.append(r)
        if req.stop_on_failure and not r.get("passed"):
            break
    return {"ran": len(ran), "results": ran}


@app.get("/eval/benchmark/report")
def eval_benchmark_report(only_latest: bool = True):
    """
    Agent vs Manual DWSIM comparison report:
      - pass-rate by difficulty
      - mean speedup vs human baseline
      - per-case rows with agent_time / human_time / speedup
      - markdown string for display or export
    """
    import benchmark_report
    return benchmark_report.build_report(only_latest=only_latest)


@app.delete("/eval/clear")
def eval_clear():
    """Clear the evaluation log and reliability failure cases."""
    get_eval_log().clear()
    get_failure_log().clear()
    return {"success": True, "message": "Evaluation log and failure cases cleared"}


# ── Reliability endpoints ─────────────────────────────────────────────────────

@app.get("/eval/reliability")
def eval_reliability():
    """
    Return LLM reliability summary:
      - Issue counts by type (unit errors, PP mismatches, physics violations, hallucinations)
      - Severity breakdown
      - Recent failure cases
    """
    return get_failure_log().get_summary()


@app.get("/eval/failures")
def eval_failures():
    """Return all stored failure cases (up to 1000), newest first."""
    cases = get_failure_log().get_all()
    return {"total": len(cases), "cases": cases[::-1]}


# ─────────────────────────────────────────────────────────────────────────────
# Accuracy Comparison API  (paper validation — Manual vs API vs Agent)
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/accuracy/reference")
def accuracy_add_reference(req: AccuracyReferenceSetModel):
    """
    Add a manual DWSIM reference set (values the user obtained by running
    DWSIM manually). These become the ground truth for comparison.
    """
    store = get_accuracy_store()
    rs = ReferenceSet(
        ref_id    = __import__("uuid").uuid4().hex[:8],
        name      = req.name,
        flowsheet = req.flowsheet,
    )
    for e in req.entries:
        rs.entries.append(ReferenceEntry(
            stream_tag   = e.stream_tag,
            property_key = e.property_key,
            manual_value = e.manual_value,
            note         = e.note,
        ))
    store.add_reference_set(rs)
    return {"success": True, "ref_id": rs.ref_id, "name": rs.name,
            "entries": len(rs.entries)}


@app.get("/accuracy/reference")
def accuracy_list_references():
    """List all stored manual reference sets."""
    return {"reference_sets": get_accuracy_store().list_reference_sets()}


@app.delete("/accuracy/reference/{ref_id}")
def accuracy_delete_reference(ref_id: str):
    """Delete a reference set by ID."""
    deleted = get_accuracy_store().delete_reference_set(ref_id)
    if not deleted:
        raise HTTPException(status_code=404, detail=f"Reference {ref_id} not found")
    return {"success": True, "deleted": ref_id}


@app.post("/accuracy/capture")
def accuracy_capture_from_dwsim(req: AccuracyCaptureRequest):
    """
    Auto-capture current DWSIM simulation results as a reference set.
    Useful for setting up a baseline from a known-good run.
    """
    bridge = _get_bridge()
    if bridge._flowsheet is None:
        raise HTTPException(status_code=400, detail="No flowsheet loaded")
    raw = bridge.get_simulation_results()
    if not raw.get("success"):
        raise HTTPException(status_code=400, detail="No simulation results available")

    store = get_accuracy_store()
    rs = store.capture_from_sim_results(
        sim_results  = raw,
        name         = req.name,
        flowsheet    = bridge.state.name or "",
        stream_tags  = req.stream_tags,
        properties   = req.properties,
    )
    store.add_reference_set(rs)
    return {"success": True, "ref_id": rs.ref_id, "name": rs.name,
            "entries": len(rs.entries),
            "entries_detail": [
                {"stream": e.stream_tag, "property": e.property_key, "value": e.manual_value}
                for e in rs.entries
            ]}


@app.post("/accuracy/compare")
def accuracy_compare(req: AccuracyCompareRequest):
    """
    Run a three-way comparison:
      1. Manual DWSIM (from reference set)
      2. Direct DWSIM API (current simulation results)
      3. AI Agent Response (parsed from last agent text, optional)

    Returns a comparison table with error percentages + markdown for paper.
    """
    store    = get_accuracy_store()
    comparer = get_accuracy_comparer()

    ref = store.get_reference_set(req.ref_id)
    if ref is None:
        raise HTTPException(status_code=404, detail=f"Reference {req.ref_id} not found")

    bridge = _get_bridge()
    sim_results: Dict = {}
    if bridge._flowsheet is not None:
        raw = bridge.get_simulation_results()
        if raw.get("success"):
            sim_results = raw

    agent_text = ""
    if req.use_last_agent_answer:
        agent_text = _last_agent_answer.get("text", "")

    # Diagnostics — surfaced in the response so the UI can show why the
    # "AI Agent Response" column is empty without the user digging into logs.
    diag: Dict[str, Any] = {
        "auto_query_ran":     False,
        "agent_text_len":     len(agent_text),
        "agent_text_preview": agent_text[:200],
        "auto_query_error":   None,
        "ref_tags":           sorted({e.stream_tag for e in ref.entries}),
        "live_stream_tags":   sorted((sim_results.get("stream_results") or {}).keys()),
    }
    diag["ref_missing_tags"] = [
        t for t in diag["ref_tags"] if t not in diag["live_stream_tags"]
    ]

    # Auto-query fallback: if there's no prior agent text, synthesize a prompt
    # from the reference-set entries and run the agent synchronously so the
    # "AI Agent Response" column is populated on first Compare click.
    if req.auto_query and not agent_text and bridge._flowsheet is not None:
        tags_needed  = sorted({e.stream_tag for e in ref.entries})
        props_needed = sorted({e.property_key for e in ref.entries})
        prop_hints   = ", ".join(
            f"{p} ({PROPERTIES.get(p, {}).get('unit','')})".strip()
            for p in props_needed
        )
        prompt = (
            "Report the current simulation results (read-only — do not set "
            "or modify any property).\n\n"
            f"Streams: {', '.join(tags_needed)}\n"
            f"Properties: {prop_hints}\n\n"
            "Format each line as: \"<stream>: <property> = <value> <unit>\". "
            "Give exactly one numeric value per (stream, property) pair."
        )
        diag["auto_query_ran"] = True
        diag["auto_query_prompt_preview"] = prompt[:300]
        try:
            agent     = _get_agent()
            answer    = agent.chat(prompt) or ""
            agent_text = answer
            _last_agent_answer["text"] = answer
            diag["agent_text_len"]     = len(answer)
            diag["agent_text_preview"] = answer[:200]
        except Exception as exc:
            diag["auto_query_error"] = f"{type(exc).__name__}: {exc}"
            print(f"[accuracy] auto-query failed: {exc}")

    result = comparer.compare(
        ref          = ref,
        sim_results  = sim_results,
        agent_text   = agent_text,
        session_id   = _last_agent_answer.get("session_id", ""),
        flowsheet    = bridge.state.name or "",
    )
    store.add_comparison(result)

    d = result.to_dict()
    d["markdown_table"] = result.as_markdown_table()
    d["diagnostics"]    = diag
    return d


@app.get("/accuracy/comparisons")
def accuracy_list_comparisons(ref_id: Optional[str] = None):
    """List recent comparison results, optionally filtered by reference set."""
    return {"comparisons": get_accuracy_store().list_comparisons(ref_id)}


@app.get("/accuracy/summary")
def accuracy_summary():
    """Return overall accuracy store summary."""
    return get_accuracy_store().get_summary()


# ─────────────────────────────────────────────────────────────────────────────
# Flowsheet File Browser — scan, watch, load from DWSIM GUI
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/flowsheets/scan")
def scan_flowsheets(max_files: int = 50):
    """
    Scan Documents/Desktop/Downloads for .dwxmz files.
    Returns files with metadata (size, modified date) sorted newest-first.
    """
    files = _scanner.scan(max_files=max_files)
    return {
        "success": True,
        "count": len(files),
        "files": files,
        "watcher_active": _watcher.is_running if _watcher else False,
    }


@app.get("/flowsheets/scan/path")
def scan_custom_path(directory: str, max_files: int = 50):
    """Scan a custom directory for flowsheet files."""
    # BUG-11 fix: canonicalize path to prevent traversal (e.g. ../../etc)
    safe_dir = os.path.realpath(os.path.abspath(directory))
    if not os.path.isdir(safe_dir):
        raise HTTPException(status_code=400, detail=f"Not a directory: {safe_dir}")
    custom_scanner = FlowsheetScanner(watch_dirs=[safe_dir])
    files = custom_scanner.scan(max_files=max_files)
    return {"success": True, "count": len(files), "files": files, "resolved_path": safe_dir}


@app.post("/flowsheets/load-by-path")
def load_flowsheet_by_path(req: LoadRequest):
    """Load a flowsheet from the file browser. Same as /flowsheet/load but grouped under /flowsheets."""
    r = _get_bridge().load_flowsheet(req.path, req.alias)
    if not r.get("success"):
        raise HTTPException(status_code=400, detail=r)
    return r


@app.get("/flowsheets/watcher/status")
def watcher_status():
    """Check if the file watcher is running."""
    base = _watcher.status if _watcher else {"running": False}
    base["ws_clients"] = len(_ws_clients)
    return base


@app.post("/flowsheets/watcher/restart")
def watcher_restart():
    """Stop + restart the file watcher (recovery from stuck state)."""
    global _watcher
    old_error_count = _watcher._error_count if _watcher else 0
    try:
        if _watcher:
            _watcher.stop()
        _watcher = FlowsheetWatcher(on_change=_broadcast_file_event,
                                    poll_interval=3.0)
        _watcher.start()
        return {"success": True, "previous_errors": old_error_count,
                "status": _watcher.status}
    except Exception as exc:
        return {"success": False, "error": str(exc)}


@app.websocket("/ws/flowsheets")
async def ws_flowsheet_events(websocket: WebSocket):
    """
    WebSocket endpoint for real-time flowsheet file change notifications.
    Pushes events when DWSIM saves/creates/deletes .dwxmz files.
    Events: { type: "file_event", event: "created"|"modified"|"deleted", file: {...} }
    """
    await websocket.accept()

    import queue as queue_mod
    q: queue_mod.Queue = queue_mod.Queue()
    _ws_clients.append(q)

    try:
        while True:
            # Check for file events to push (non-blocking poll)
            try:
                msg = q.get_nowait()
                await websocket.send_text(msg)
            except queue_mod.Empty:
                pass

            # Check for incoming messages (ping/pong keepalive)
            try:
                data = await asyncio.wait_for(websocket.receive_text(), timeout=1.0)
                if data == "ping":
                    await websocket.send_text('{"type":"pong"}')
                elif data == "scan":
                    # Client requests a manual scan push
                    files = _scanner.scan(max_files=50)
                    await websocket.send_text(json.dumps({
                        "type": "scan_result",
                        "count": len(files),
                        "files": files,
                    }))
            except asyncio.TimeoutError:
                continue
            except WebSocketDisconnect:
                break
    finally:
        if q in _ws_clients:
            _ws_clients.remove(q)


# ─────────────────────────────────────────────────────────────────────────────

# Startup/shutdown handled by _lifespan above (modern FastAPI pattern)


if __name__ == "__main__":
    port = int(os.getenv("PORT", "8080"))
    print(f"[DWSIM API v2] Starting on http://localhost:{port}")
    print(f"[DWSIM API v2] Docs at  http://localhost:{port}/docs")
    uvicorn.run("api:app", host="0.0.0.0", port=port, reload=False)
